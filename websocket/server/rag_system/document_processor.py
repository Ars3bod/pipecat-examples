#
# Copyright (c) 2025, Daily
#
# SPDX-License-Identifier: BSD 2-Clause License
#

"""
Document Processing Pipeline for RAG Knowledge Management System
"""

import os
import hashlib
import logging
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime
import uuid

# Document processing libraries
import PyPDF2
from docx import Document
import openpyxl
import pandas as pd
from langdetect import detect, LangDetectException

# Text processing
import re
import nltk
from nltk.tokenize import sent_tokenize, word_tokenize
from nltk.corpus import stopwords

# Configuration
from .config import get_config, get_document_config, get_metadata_schema, VALIDATION_RULES

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class DocumentProcessor:
    """
    Handles document processing for the RAG knowledge management system.
    """
    
    def __init__(self):
        self.config = get_config()
        self.doc_config = get_document_config()
        self.metadata_schema = get_metadata_schema()
        self.validation_rules = VALIDATION_RULES
        
        # Create directories if they don't exist
        self._create_directories()
        
        # Download required NLTK data
        self._setup_nltk()
    
    def _create_directories(self):
        """Create necessary directories."""
        directories = [
            self.config["documents"]["upload_directory"],
            self.config["documents"]["processed_directory"],
            self.config["documents"]["metadata_directory"]
        ]
        
        for directory in directories:
            os.makedirs(directory, exist_ok=True)
    
    def _setup_nltk(self):
        """Download required NLTK data."""
        try:
            nltk.data.find('tokenizers/punkt')
        except LookupError:
            nltk.download('punkt')
        
        try:
            nltk.data.find('corpora/stopwords')
        except LookupError:
            nltk.download('stopwords')
    
    def process_document(self, file_path: str, metadata: Dict[str, Any]) -> Dict[str, Any]:
        """
        Process a document and return chunks with metadata.
        
        Args:
            file_path: Path to the document file
            metadata: Document metadata
            
        Returns:
            Dictionary containing processed chunks and metadata
        """
        try:
            # Validate file
            if not self._validate_file(file_path):
                raise ValueError(f"Invalid file: {file_path}")
            
            # Extract text
            text = self.extract_text(file_path)
            if not text.strip():
                raise ValueError(f"No text content found in file: {file_path}")
            
            # Detect language
            language = self.detect_language(text)
            
            # Clean text
            cleaned_text = self.clean_text(text, language)
            
            # Chunk text
            chunks = self.chunk_text(cleaned_text)
            
            # Generate document ID and metadata
            document_id = self._generate_document_id(file_path, metadata)
            file_hash = self._calculate_file_hash(file_path)
            
            # Prepare result
            result = {
                "document_id": document_id,
                "file_path": file_path,
                "file_hash": file_hash,
                "language": language,
                "chunks": chunks,
                "metadata": {
                    **metadata,
                    "document_id": document_id,
                    "file_hash": file_hash,
                    "language": language,
                    "chunk_count": len(chunks),
                    "processed_date": datetime.now().isoformat(),
                    "file_size": os.path.getsize(file_path)
                }
            }
            
            logger.info(f"Successfully processed document: {document_id}")
            return result
            
        except Exception as e:
            logger.error(f"Error processing document {file_path}: {str(e)}")
            raise
    
    def extract_text(self, file_path: str) -> str:
        """
        Extract text from various document formats.
        
        Args:
            file_path: Path to the document file
            
        Returns:
            Extracted text content
        """
        file_extension = os.path.splitext(file_path)[1].lower()
        
        if file_extension == '.pdf':
            return self._extract_pdf_text(file_path)
        elif file_extension == '.docx':
            return self._extract_docx_text(file_path)
        elif file_extension == '.xlsx':
            return self._extract_xlsx_text(file_path)
        elif file_extension == '.txt':
            return self._extract_txt_text(file_path)
        elif file_extension == '.md':
            return self._extract_md_text(file_path)
        else:
            raise ValueError(f"Unsupported file format: {file_extension}")
    
    def _extract_pdf_text(self, file_path: str) -> str:
        """Extract text from PDF file."""
        text = ""
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
        except Exception as e:
            logger.error(f"Error extracting PDF text: {str(e)}")
            raise
        return text
    
    def _extract_docx_text(self, file_path: str) -> str:
        """Extract text from DOCX file."""
        text = ""
        try:
            doc = Document(file_path)
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
        except Exception as e:
            logger.error(f"Error extracting DOCX text: {str(e)}")
            raise
        return text
    
    def _extract_xlsx_text(self, file_path: str) -> str:
        """Extract text from XLSX file."""
        text = ""
        try:
            workbook = openpyxl.load_workbook(file_path)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(values_only=True):
                    row_text = " ".join([str(cell) for cell in row if cell is not None])
                    if row_text.strip():
                        text += row_text + "\n"
        except Exception as e:
            logger.error(f"Error extracting XLSX text: {str(e)}")
            raise
        return text
    
    def _extract_txt_text(self, file_path: str) -> str:
        """Extract text from TXT file."""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                return file.read()
        except UnicodeDecodeError:
            # Try with different encodings
            encodings = ['latin-1', 'cp1252', 'iso-8859-1']
            for encoding in encodings:
                try:
                    with open(file_path, 'r', encoding=encoding) as file:
                        return file.read()
                except UnicodeDecodeError:
                    continue
            raise ValueError(f"Could not decode file with any supported encoding: {file_path}")
    
    def _extract_md_text(self, file_path: str) -> str:
        """Extract text from Markdown file."""
        return self._extract_txt_text(file_path)  # Same as TXT for now
    
    def detect_language(self, text: str) -> str:
        """
        Detect the language of the text.
        
        Args:
            text: Text content
            
        Returns:
            Language code ('ar' or 'en')
        """
        try:
            # Use first 1000 characters for detection
            sample_text = text[:1000]
            detected_lang = detect(sample_text)
            
            # Map to supported languages
            if detected_lang in ['ar', 'arabic']:
                return 'ar'
            else:
                return 'en'
                
        except LangDetectException:
            # Default to Arabic if detection fails
            return 'ar'
    
    def clean_text(self, text: str, language: str) -> str:
        """
        Clean and normalize text content.
        
        Args:
            text: Raw text content
            language: Language code
            
        Returns:
            Cleaned text
        """
        # Remove extra whitespace
        text = re.sub(r'\s+', ' ', text)
        
        # Remove special characters but keep Arabic/English text
        if language == 'ar':
            # Keep Arabic characters, numbers, and basic punctuation
            text = re.sub(r'[^\u0600-\u06FF\u0750-\u077F\u08A0-\u08FF\uFB50-\uFDFF\uFE70-\uFEFF\s\d\.\,\!\?\:\;\(\)]', '', text)
        else:
            # Keep English characters, numbers, and basic punctuation
            text = re.sub(r'[^a-zA-Z0-9\s\.\,\!\?\:\;\(\)]', '', text)
        
        # Normalize whitespace
        text = re.sub(r'\s+', ' ', text).strip()
        
        return text
    
    def chunk_text(self, text: str) -> List[Dict[str, Any]]:
        """
        Split text into overlapping chunks.
        
        Args:
            text: Cleaned text content
            
        Returns:
            List of text chunks with metadata
        """
        chunk_size = self.config["chunking"]["chunk_size"]
        chunk_overlap = self.config["chunking"]["chunk_overlap"]
        min_chunk_size = self.config["chunking"]["min_chunk_size"]
        separators = self.config["chunking"]["separators"]
        
        chunks = []
        
        # Split by sentences first
        sentences = sent_tokenize(text)
        
        current_chunk = ""
        chunk_index = 0
        
        for sentence in sentences:
            # Check if adding this sentence would exceed chunk size
            if len(current_chunk) + len(sentence) > chunk_size and current_chunk:
                # Save current chunk
                if len(current_chunk.strip()) >= min_chunk_size:
                    chunks.append({
                        "chunk_id": f"{chunk_index}",
                        "content": current_chunk.strip(),
                        "chunk_index": chunk_index,
                        "start_char": len(" ".join(chunks)) if chunks else 0,
                        "end_char": len(" ".join(chunks)) + len(current_chunk.strip())
                    })
                    chunk_index += 1
                
                # Start new chunk with overlap
                overlap_text = current_chunk[-chunk_overlap:] if chunk_overlap > 0 else ""
                current_chunk = overlap_text + " " + sentence
            else:
                current_chunk += " " + sentence if current_chunk else sentence
        
        # Add final chunk
        if current_chunk.strip() and len(current_chunk.strip()) >= min_chunk_size:
            chunks.append({
                "chunk_id": f"{chunk_index}",
                "content": current_chunk.strip(),
                "chunk_index": chunk_index,
                "start_char": len(" ".join([c["content"] for c in chunks])),
                "end_char": len(" ".join([c["content"] for c in chunks])) + len(current_chunk.strip())
            })
        
        return chunks
    
    def _validate_file(self, file_path: str) -> bool:
        """Validate file before processing."""
        if not os.path.exists(file_path):
            return False
        
        file_size = os.path.getsize(file_path)
        if file_size > self.validation_rules["file_size"]["max"]:
            return False
        
        if file_size < self.validation_rules["file_size"]["min"]:
            return False
        
        file_extension = os.path.splitext(file_path)[1].lower()
        if file_extension not in self.validation_rules["file_types"]:
            return False
        
        return True
    
    def _generate_document_id(self, file_path: str, metadata: Dict[str, Any]) -> str:
        """Generate unique document ID."""
        # Use filename and metadata to generate ID
        filename = os.path.basename(file_path)
        title = metadata.get("title", filename)
        department = metadata.get("department", "unknown")
        
        # Create hash from filename, title, and department
        hash_input = f"{filename}_{title}_{department}_{datetime.now().isoformat()}"
        return hashlib.md5(hash_input.encode()).hexdigest()[:16]
    
    def _calculate_file_hash(self, file_path: str) -> str:
        """Calculate SHA256 hash of file."""
        hash_sha256 = hashlib.sha256()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_sha256.update(chunk)
        return hash_sha256.hexdigest()

# Example usage
if __name__ == "__main__":
    processor = DocumentProcessor()
    
    # Example metadata
    metadata = {
        "title": "HR Policy Manual",
        "department": "HR",
        "category": "policies",
        "author": "HR Department",
        "version": "1.0"
    }
    
    # Process a document (replace with actual file path)
    # result = processor.process_document("sample_document.pdf", metadata)
    # print(f"Processed {len(result['chunks'])} chunks")
    # print(f"Language: {result['language']}")
    # print(f"Document ID: {result['document_id']}")
